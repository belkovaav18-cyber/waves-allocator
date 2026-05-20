import streamlit as st
import pandas as pd
from datetime import datetime
import re
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from sheets import load_guests, save_results_with_details, load_registration_data
from preprocess import preprocess_guests
from solver_controller import smart_solve
from comment_engine import process_comments, display_comments_report
from feasibility import check_feasibility, display_feasibility_report

# =========================================================
# КОНСТАНТЫ
# =========================================================
SHEET_ID = "1lF4SV24wTo5OwsidQ7UPqBVaGzdw_fBSx0OuBJJ4cWg"
REGISTRATION_SHEET_ID = "1fHjI0hTtlbjDZxSCWVidzGnJ7aY4joB7UUVXFEB0rxw"
TAB_NAME = "Sheet"

st.set_page_config(page_title="Система расселения", layout="wide")
st.title("🏨 Система расселения в Доме отдыха")

# Инициализация session state
if 'layout' not in st.session_state:
    st.session_state.layout = None
if 'final_result_df' not in st.session_state:
    st.session_state.final_result_df = None
if 'data_loaded' not in st.session_state:
    st.session_state.data_loaded = False
if 'edit_mode' not in st.session_state:
    st.session_state.edit_mode = False

# Загрузка комнат
try:
    rooms_df = pd.read_excel("data/rooms.xlsx")
    rooms = rooms_df.to_dict("records")
    # Создаем список всех комнат для выпадающего списка
    all_rooms_list = sorted([str(r['room_id']) for r in rooms])
except Exception as e:
    st.error(f"❌ Ошибка загрузки комнат: {e}")
    st.stop()

# =========================================================
# ЗАГРУЗКА ДАННЫХ
# =========================================================
if not st.session_state.data_loaded:
    with st.spinner("Загрузка данных из Google Sheets..."):
        raw = load_guests(SHEET_ID, TAB_NAME)
        registration_df = load_registration_data(REGISTRATION_SHEET_ID, TAB_NAME)
        
        if raw.empty:
            st.error("❌ Не удалось загрузить данные гостей")
            st.stop()
        
        guests_df = preprocess_guests(raw, registration_df)
        
        st.session_state.raw = raw
        st.session_state.registration_df = registration_df
        st.session_state.guests_df = guests_df
        st.session_state.data_loaded = True
        
        # Отладочная информация
        residents_count = len(guests_df[guests_df['resident'] == True])
        non_residents_count = len(guests_df[guests_df['resident'] == False])
        
        with st.expander("🔍 Отладочная информация о резидентах"):
            st.write(f"Всего гостей: {len(guests_df)}")
            st.write(f"Резиденты (будут жить): {residents_count}")
            st.write(f"Нерезиденты (не будут жить): {non_residents_count}")
        
        st.success(f"✅ Загружено {len(guests_df)} гостей (из них резидентов: {residents_count})")
else:
    raw = st.session_state.raw
    registration_df = st.session_state.registration_df
    guests_df = st.session_state.guests_df


# =========================================================
# ФУНКЦИИ ДЛЯ РАСЧЕТА ДАТ
# =========================================================
def extract_dates_from_guest(guest_fio, raw_df):
    month_map = {
        "мая": 5, "июня": 6, "июля": 7, "августа": 8,
        "сентября": 9, "октября": 10, "ноября": 11, "декабря": 12,
        "января": 1, "февраля": 2, "марта": 3, "апреля": 4
    }
    
    guest_row = None
    for idx, row in raw_df.iterrows():
        if row.get("ФИО") == guest_fio:
            guest_row = row
            break
    
    if guest_row is None:
        return "", ""
    
    selected_nights = []
    for col in raw_df.columns:
        col_str = str(col)
        if "Комната" in col_str and "ночь" in col_str:
            value = guest_row.get(col, "")
            if pd.notna(value) and str(value).strip().lower() not in ["", "нет", "-", "false", "nan"]:
                selected_nights.append(col_str)
    
    if not selected_nights:
        return "", ""
    
    def get_date_from_column(col_str):
        match = re.search(r"ночь на (\d+) (\w+)", col_str)
        if match:
            day = int(match.group(1))
            month_name = match.group(2)
            month = month_map.get(month_name, 6)
            year = 2026
            return (month, day, datetime(year, month, day))
        return (99, 99, None)
    
    nights_with_dates = []
    for night in selected_nights:
        month, day, date = get_date_from_column(night)
        if date:
            nights_with_dates.append((date, night, month, day))
    
    nights_with_dates.sort(key=lambda x: x[0])
    
    if not nights_with_dates:
        return "", ""
    
    first_date, first_night, first_month, first_day = nights_with_dates[0]
    
    if first_month == 6 and first_day == 1:
        check_in = datetime(2026, 5, 31)
    else:
        check_in = datetime(2026, first_month, first_day - 1)
    check_in_str = check_in.strftime("%d.%m.%Y")
    
    last_date, last_night, last_month, last_day = nights_with_dates[-1]
    check_out = datetime(2026, last_month, last_day)
    check_out_str = check_out.strftime("%d.%m.%Y")
    
    return check_in_str, check_out_str


# =========================================================
# ФУНКЦИИ ДЛЯ ВИЗУАЛИЗАЦИИ ПЛАНА ЭТАЖЕЙ
# =========================================================
def get_building_and_floor(room_id):
    room_str = str(room_id)
    if '-' in room_str:
        parts = room_str.split('-')
        building = parts[0]
        room_num = parts[1]
        floor = room_num[0] if room_num else '0'
        return building, floor
    else:
        floor = room_str[0] if room_str else '0'
        return None, floor


def create_floor_layout(allocation_df, rooms_df):
    allocation_dict = {}
    for _, row in allocation_df.iterrows():
        room_id = row.get('room_id')
        fio = row.get('ФИО', row.get('fio', ''))
        if room_id and fio and room_id != "не проживает" and "требуется" not in str(room_id):
            if room_id not in allocation_dict:
                allocation_dict[room_id] = []
            allocation_dict[room_id].append(fio)
    
    layout = {}
    
    for _, room in rooms_df.iterrows():
        room_id = room.get('room_id')
        capacity = room.get('вместимость', 1)
        
        building, floor = get_building_and_floor(room_id)
        
        if building is None or building == 'None':
            if str(room_id).startswith('1'):
                building = '1'
            elif str(room_id).startswith('2'):
                building = '2'
            else:
                building = 'unknown'
        
        if building not in layout:
            layout[building] = {}
        if floor not in layout[building]:
            layout[building][floor] = {}
        
        layout[building][floor][room_id] = {
            'capacity': capacity,
            'guests': allocation_dict.get(room_id, []),
            'free_spots': capacity - len(allocation_dict.get(room_id, []))
        }
    
    return layout


def render_floor_plan(layout, selected_building=None):
    if selected_building:
        buildings = [selected_building]
    else:
        buildings = sorted([b for b in layout.keys() if b != 'unknown'])
        if 'unknown' in layout:
            buildings.append('unknown')
    
    for building in buildings:
        if building not in layout:
            continue
        
        building_name = "Красный" if building == "1" else ("Желтый" if building == "2" else building)
        st.markdown(f"### 🏢 Корпус {building_name}")
        
        floors = sorted(layout[building].keys(), key=int)
        
        for floor in floors:
            st.markdown(f"#### 📍 Этаж {floor}")
            
            rooms = layout[building][floor]
            sorted_rooms = sorted(rooms.items(), key=lambda x: int(re.sub(r'\D', '', x[0])))
            
            cols = st.columns(4)
            
            for idx, (room_id, room_data) in enumerate(sorted_rooms):
                col = cols[idx % 4]
                
                if room_data['free_spots'] == 0:
                    status_color = "#28a745"
                    status_text = "✅ Полная"
                elif room_data['free_spots'] == room_data['capacity']:
                    status_color = "#dc3545"
                    status_text = "❌ Пустая"
                else:
                    status_color = "#ffc107"
                    status_text = f"⚠️ Свободно: {room_data['free_spots']}"
                
                guests_list = "<br>".join(room_data['guests']) if room_data['guests'] else "—"
                
                with col:
                    st.markdown(
                        f"""
                        <div style="
                            border: 2px solid {status_color};
                            border-radius: 10px;
                            padding: 10px;
                            margin: 5px;
                            background-color: #f8f9fa;
                        ">
                            <h3 style="text-align: center; margin: 0; color: {status_color};">
                                {room_id}
                            </h3>
                            <p style="text-align: center; margin: 5px 0;">
                                🛏️ {room_data['capacity']} {'место' if room_data['capacity'] == 1 else 'места'}
                            </p>
                            <p style="text-align: center; margin: 5px 0;">
                                📊 {status_text}
                            </p>
                            <hr style="margin: 5px 0;">
                            <div style="font-size: 12px;">
                                <strong>👥 Заселены:</strong><br>
                                {guests_list}
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
            
            st.markdown("---")


# =========================================================
# ФУНКЦИЯ ЭКСПОРТА В EXCEL
# =========================================================
def export_to_excel_with_styles(layout):
    """Экспортирует план этажей в Excel с цветами ячеек"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for building in layout:
            if building == 'unknown':
                continue
            
            building_name = "Красный_корпус" if building == "1" else "Желтый_корпус" if building == "2" else f"Корпус_{building}"
            
            floors = sorted(layout[building].keys(), key=int)
            
            for floor in floors:
                sheet_name = f"{building_name}_этаж_{floor}"[:31]
                
                rooms = layout[building][floor]
                sorted_rooms = sorted(rooms.items(), key=lambda x: int(re.sub(r'\D', '', x[0])))
                
                data = []
                row_data = []
                max_rooms_in_row = 4
                rooms_info = {}
                
                for idx, (room_id, room_data) in enumerate(sorted_rooms):
                    guests_text = "\n".join(room_data['guests']) if room_data['guests'] else "—"
                    cell_text = f"{room_id}\nМест: {room_data['capacity']}\nСвободно: {room_data['free_spots']}\n\nЗаселены:\n{guests_text}"
                    row_data.append(cell_text)
                    rooms_info[(idx // max_rooms_in_row, idx % max_rooms_in_row)] = room_data
                    
                    if len(row_data) == max_rooms_in_row:
                        data.append(row_data)
                        row_data = []
                
                if row_data:
                    while len(row_data) < max_rooms_in_row:
                        row_data.append("")
                    data.append(row_data)
                
                df = pd.DataFrame(data)
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                green_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
                red_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
                
                center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row_idx, row in enumerate(data, start=1):
                    for col_idx, cell_value in enumerate(row, start=1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        room_key = (row_idx - 1, col_idx - 1)
                        if room_key in rooms_info:
                            room_data = rooms_info[room_key]
                            if room_data['free_spots'] == 0:
                                cell.fill = green_fill
                                cell.font = Font(color="FFFFFF", bold=True)
                            elif room_data['free_spots'] == room_data['capacity']:
                                cell.fill = red_fill
                                cell.font = Font(color="FFFFFF", bold=True)
                            else:
                                cell.fill = yellow_fill
                                cell.font = Font(color="000000", bold=True)
                        else:
                            cell.fill = red_fill
                            cell.font = Font(color="FFFFFF", bold=True)
                        
                        cell.alignment = center_alignment
                        cell.border = border
                        col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + (col_idx // 26)) + chr(64 + (col_idx % 26))
                        worksheet.column_dimensions[col_letter].width = 30
                        worksheet.row_dimensions[row_idx].height = 120
    
    output.seek(0)
    return output


# =========================================================
# ФУНКЦИЯ ОБНОВЛЕНИЯ РАССЕЛЕНИЯ
# =========================================================
def update_allocation():
    """Обновляет план этажей после ручного изменения"""
    if st.session_state.final_result_df is not None:
        allocated_guests = st.session_state.final_result_df[
            st.session_state.final_result_df["room_id"].apply(lambda x: str(x) not in ['не проживает', 'нет мест', 'требуется ручная обработка'])
        ]
        if len(allocated_guests) > 0:
            st.session_state.layout = create_floor_layout(allocated_guests, rooms_df)
        
        # Сохраняем в Google Sheets
        try:
            save_results_with_details(
                SHEET_ID, 
                "Result", 
                st.session_state.final_result_df, 
                raw,
                guests_df
            )
            st.success("✅ Изменения сохранены в Google Sheets")
        except Exception as e:
            st.warning(f"Не удалось сохранить в Google Sheets: {e}")


# =========================================================
# РЕДАКТОР РАССЕЛЕНИЯ
# =========================================================
def manual_allocation_editor():
    """Интерфейс для ручного расселения с отображением жильцов в комнатах"""
    st.subheader("✏️ Ручное расселение")
    
    if st.session_state.final_result_df is None:
        st.warning("Сначала выполните автоматическое расселение")
        return
    
    edit_df = st.session_state.final_result_df.copy()
    fio_col = 'ФИО' if 'ФИО' in edit_df.columns else 'fio' if 'fio' in edit_df.columns else None
    
    if fio_col is None:
        st.error("Не найдена колонка с ФИО")
        return
    
    # Фильтруем резидентов
    residents_only = edit_df[edit_df["room_id"] != "не проживает"].copy()
    
    if len(residents_only) == 0:
        st.info("Нет резидентов для ручного расселения")
        return
    
    # Получаем актуальную информацию о комнатах из Google Sheets
    try:
        sheet = get_google_client().open_by_key(SHEET_ID)
        result_sheet = sheet.worksheet("Result")
        result_data = result_sheet.get_all_records()
        current_df = pd.DataFrame(result_data)
        if not current_df.empty and 'room_id' in current_df.columns:
            # Обновляем локальную копию
            for idx, row in residents_only.iterrows():
                fio = row[fio_col]
                current_row = current_df[current_df[fio_col] == fio]
                if not current_row.empty:
                    residents_only.loc[idx, 'room_id'] = current_row.iloc[0].get('room_id', row['room_id'])
    except:
        pass
    
    # Создаем словарь жильцов по комнатам
    room_occupants = {}
    for _, row in residents_only.iterrows():
        room_id = row['room_id']
        fio = row[fio_col]
        if room_id not in room_occupants:
            room_occupants[room_id] = []
        room_occupants[room_id].append(fio)
    
    # Получаем вместимость комнат
    room_capacity = {r['room_id']: r['вместимость'] for r in rooms}
    
    # Поиск гостя
    search_term = st.text_input("🔍 Поиск гостя по фамилии", placeholder="Введите фамилию...")
    
    filtered_guests = residents_only
    if search_term:
        filtered_guests = residents_only[residents_only[fio_col].str.contains(search_term, case=False, na=False)]
    
    st.write(f"**Найдено гостей: {len(filtered_guests)}**")
    
    # Для каждого гостя
    for idx, (_, guest) in enumerate(filtered_guests.iterrows()):
        col1, col2, col3 = st.columns([3, 3, 1])
        
        with col1:
            st.write(f"**{guest[fio_col]}**")
            if guest.get('comment') and str(guest['comment']) != 'nan':
                st.caption(f"💬 {guest['comment'][:100]}...")
        
        with col2:
            current_room = guest['room_id']
            st.write(f"Текущая: **{current_room}** (вместимость: {room_capacity.get(current_room, '?')})")
            
            # Формируем список комнат с информацией о жильцах
            room_options = ["--- оставить как есть ---"]
            
            # Сортируем комнаты по корпусу и номеру
            sorted_rooms = sorted(rooms, key=lambda x: x['room_id'])
            
            for room in sorted_rooms:
                room_id = room['room_id']
                capacity = room['вместимость']
                occupants = room_occupants.get(room_id, [])
                free_spots = capacity - len(occupants)
                
                # Формируем описание
                if room_id == current_room:
                    status = "📍 текущая"
                elif free_spots > 0:
                    status = f"🟢 свободно {free_spots}/{capacity}"
                else:
                    status = "🔴 занято"
                
                # Список жильцов
                if occupants:
                    occupants_str = ", ".join([o.split()[0] for o in occupants[:2]])  # Только фамилии
                    if len(occupants) > 2:
                        occupants_str += f" +{len(occupants)-2}"
                    room_label = f"{room_id} [{status}] 👥 {occupants_str}"
                else:
                    room_label = f"{room_id} [{status}] 🏠 пусто"
                
                room_options.append(room_label)
            
            room_options.append("❌ не проживает")
            
            selected = st.selectbox(
                "Новая комната",
                options=room_options,
                index=0,
                key=f"room_select_{idx}_{guest[fio_col]}",
                label_visibility="collapsed"
            )
            
            if selected != "--- оставить как есть ---":
                if selected == "❌ не проживает":
                    new_room = "не проживает"
                else:
                    new_room = selected.split("[")[0].strip()
                
                # Обновляем в session_state
                mask = st.session_state.final_result_df[fio_col] == guest[fio_col]
                st.session_state.final_result_df.loc[mask, 'room_id'] = new_room
                
                if new_room != "не проживает":
                    room_info = next((r for r in rooms if r['room_id'] == new_room), None)
                    if room_info:
                        st.session_state.final_result_df.loc[mask, 'room_capacity'] = room_info['вместимость']
                
                # Обновляем локальный словарь жильцов
                old_room = current_room
                if old_room in room_occupants and guest[fio_col] in room_occupants[old_room]:
                    room_occupants[old_room].remove(guest[fio_col])
                if new_room != "не проживает":
                    if new_room not in room_occupants:
                        room_occupants[new_room] = []
                    room_occupants[new_room].append(guest[fio_col])
                
                st.rerun()
        
        with col3:
            if st.button("🔄", key=f"refresh_{idx}_{guest[fio_col]}"):
                update_allocation()
                st.rerun()
        
        st.divider()
    
    # Кнопки действий
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        if st.button("💾 Сохранить все изменения", type="primary"):
            update_allocation()
            st.success("Изменения сохранены!")
            st.rerun()
    
    with col2:
        if st.button("🔄 Обновить план этажей"):
            update_allocation()
            st.rerun()
    
    with col3:
        if st.button("📊 Показать статистику"):
            stats = st.session_state.final_result_df[
                st.session_state.final_result_df["room_id"] != "не проживает"
            ]['room_id'].value_counts()
            st.write("### 📊 Статистика заселения")
            st.write(f"**Всего занято комнат:** {len(stats)}")
            st.write(f"**Среднее количество человек в комнате:** {stats.mean():.1f}")
            st.dataframe(stats.head(20))
    
    with col4:
        # Показать свободные места
        if st.button("🏠 Свободные места"):
            st.write("### 🏠 Комнаты со свободными местами")
            for room_id, occupants in room_occupants.items():
                capacity = room_capacity.get(room_id, 2)
                free = capacity - len(occupants)
                if free > 0:
                    st.write(f"**{room_id}**: свободно {free}/{capacity}, жильцы: {', '.join(occupants) if occupants else 'нет'}")


# =========================================================
# ОСНОВНОЙ КОД
# =========================================================
# Разделяем на резидентов и нерезидентов
non_residents = guests_df[guests_df["resident"] == False].copy()
residents = guests_df[guests_df["resident"] == True].copy()

# Отображаем информацию
col1, col2, col3 = st.columns(3)
with col1:
    st.metric("Всего гостей", len(guests_df))
with col2:
    st.metric("Резиденты", len(residents))
with col3:
    st.metric("Нерезиденты", len(non_residents))

# Проверка возможности расселения
issues = check_feasibility(guests_df, rooms_df)
display_feasibility_report(issues)

# Отображение комментариев
comments_report = process_comments(guests_df)
display_comments_report(comments_report)

st.markdown("---")

# Вкладки для автоматического и ручного расселения
tab1, tab2, tab3 = st.tabs(["🚀 Автоматическое расселение", "✏️ Ручное расселение", "📋 Список гостей"])

with tab1:
    # Таблица с гостями
    st.subheader("📋 Список резидентов")
    st.dataframe(residents[['ФИО', 'возраст', 'должность', 'город', 'comment']], width="stretch")
    
    # Кнопка расселения
    if st.button("🚀 Запустить автоматическое расселение", type="primary", width="stretch"):
        with st.spinner("Идет расселение..."):
            result, debug = smart_solve(
                residents.to_dict("records"),
                rooms
            )
        
        if "error" in result.columns:
            st.error(f"❌ Ошибка при расселении: {result.iloc[0]['error']}")
        elif len(result) == 0:
            st.error("❌ Не удалось найти решение для расселения")
        else:
            # Добавляем даты
            result_with_dates = []
            for _, row in result.iterrows():
                guest_data = row.to_dict()
                if 'ФИО' not in guest_data and 'fio' in guest_data:
                    guest_data['ФИО'] = guest_data['fio']
                
                if 'ФИО' in guest_data:
                    check_in, check_out = extract_dates_from_guest(guest_data["ФИО"], raw)
                    guest_data["Дата заезда"] = check_in
                    guest_data["Дата отъезда"] = check_out
                    result_with_dates.append(guest_data)
            
            result_df = pd.DataFrame(result_with_dates)
            
            # Нерезиденты
            non_residents_with_dates = []
            for _, row in non_residents.iterrows():
                guest_data = row.to_dict()
                check_in, check_out = extract_dates_from_guest(guest_data["ФИО"], raw)
                guest_data["Дата заезда"] = check_in
                guest_data["Дата отъезда"] = check_out
                guest_data["room_id"] = "не проживает"
                guest_data["room_capacity"] = 0
                non_residents_with_dates.append(guest_data)
            
            non_residents_df = pd.DataFrame(non_residents_with_dates)
            
            # Объединяем
            final_result = pd.concat([result_df, non_residents_df], ignore_index=True)
            st.session_state.final_result_df = final_result
            
            # Создаем layout
            allocated_guests = final_result[final_result["room_id"].apply(lambda x: str(x) not in ['не проживает', 'нет мест', 'требуется ручная обработка'])]
            if len(allocated_guests) > 0:
                st.session_state.layout = create_floor_layout(allocated_guests, rooms_df)
            
            # Сохраняем в Google Sheets
            try:
                save_results_with_details(SHEET_ID, "Result", final_result, raw, guests_df)
                st.success("✅ Расселение выполнено! Результат сохранен в Google Sheets")
            except Exception as e:
                st.warning(f"Не удалось сохранить в Google Sheets: {e}")
            
            st.rerun()

with tab2:
    manual_allocation_editor()
    
    # Отображение плана этажей
    if st.session_state.layout:
        st.markdown("---")
        st.subheader("🏠 План этажей с расселением")
        
        building_filter = st.radio(
            "Выберите корпус:",
            options=["Все", "Красный (1)", "Желтый (2)"],
            horizontal=True,
            key="building_filter_manual"
        )
        
        if building_filter == "Красный (1)":
            selected_building = "1"
        elif building_filter == "Желтый (2)":
            selected_building = "2"
        else:
            selected_building = None
        
        render_floor_plan(st.session_state.layout, selected_building)

with tab3:
    st.subheader("📋 Полный список гостей")
    st.dataframe(guests_df, width="stretch")
    
    # Экспорт в CSV
    csv = guests_df.to_csv(index=False).encode('utf-8-sig')
    st.download_button(
        label="📥 Скачать список гостей (CSV)",
        data=csv,
        file_name=f"guests_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
        mime="text/csv"
    )


# =========================================================
# ОТОБРАЖЕНИЕ РЕЗУЛЬТАТОВ ПОСЛЕ РАССЕЛЕНИЯ
# =========================================================
if st.session_state.final_result_df is not None and st.session_state.layout:
    st.markdown("---")
    st.subheader("📊 Текущее расселение")
    
    # Таблица расселения
    fio_col = 'ФИО' if 'ФИО' in st.session_state.final_result_df.columns else 'fio'
    display_cols = [fio_col, 'возраст', 'должность', 'город', 'room_id', 'room_capacity', 'Дата заезда', 'Дата отъезда']
    existing_cols = [c for c in display_cols if c in st.session_state.final_result_df.columns]
    st.dataframe(st.session_state.final_result_df[existing_cols], width="stretch")
    
    # Экспорт
    col1, col2 = st.columns(2)
    with col1:
        if st.button("📑 Экспорт плана этажей в Excel"):
            try:
                excel_bytes = export_to_excel_with_styles(st.session_state.layout)
                st.download_button(
                    label="📥 Скачать",
                    data=excel_bytes,
                    file_name=f"plan_raseleniya_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Ошибка: {e}")
    
    with col2:
        csv = st.session_state.final_result_df.to_csv(index=False).encode('utf-8-sig')
        st.download_button(
            label="📥 Скачать таблицу (CSV)",
            data=csv,
            file_name=f"raselenie_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv"
        )
